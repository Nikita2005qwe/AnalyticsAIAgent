"""
Версия: v0.1
Автор: Боряков
Дата: 21.08.2025
Статус: полностью готово
Помогает, пока что прочитать накладные из файла excel.
В будущем будет полноценным оркестратором для функций обработки и чтения файлов разных форматов.
"""
import os
from typing import List, Dict
import pandas as pd
from openpyxl import load_workbook

from src.core.logger import Logger


class FileHandler:
    """
    Обработчик Excel-файла.
    Читает файл, начиная с 12-й строки (индекс 10), считая 11-ю строку заголовками.
    """

    def __init__(self, file_path: str, logger: Logger):
        self.file_path = file_path
        self.logger = logger
        self.full_path = os.path.join(os.getcwd(), file_path)

    def read_invoices(self, sheet_name: str) -> List[Dict]:
        """
        Читает Excel-файл и возвращает список словарей с данными о накладных.

        Строки:
            - Строка 1–10: фильтры, сводки — пропускаются.
            - Строка 11: заголовки (столбцы).
            - Строка 12 и далее: данные.

        Returns:
            List[Dict]: Список строк в виде словарей.
        """
        if not os.path.exists(self.full_path):
            raise FileNotFoundError(f"Файл не найден: {self.full_path}")

        try:
            # Пропускаем первые 10 строк (0-индексация: строки 0–9), 10-я — заголовки (индекс 10)
            df = pd.read_excel(self.full_path, sheet_name=sheet_name, skiprows=10, header=0, dtype=str)

            # Убедимся, что у нас есть нужные столбцы
            expected_columns = {"Названия строк", "ID XCRM Parent", "Address Normalized", "ISA", "SFA"}
            missing = expected_columns - set(df.columns)
            if missing:
                raise ValueError(f"Отсутствуют обязательные столбцы: {missing}")

            # Переименовываем только нужные столбцы
            data = df.rename(columns={
                "Названия строк": "number",
                "ID XCRM Parent": "crm_id",
                "Address Normalized": "address",
                "ISA": "isa_amount",
                "SFA": "sfa_amount",
            })

            # Оставляем только нужные колонки
            data = data[["number", "crm_id", "address", "isa_amount", "sfa_amount"]]

            # Конвертируем в список словарей
            records = data.to_dict(orient="records")

            # Приводим пустые строки к None
            cleaned_records = []
            for record in records:
                cleaned = {
                    k: (v.strip() if isinstance(v, str) and v.strip() != "" else None)
                    for k, v in record.items()
                }
                cleaned_records.append(cleaned)

            return cleaned_records

        except ValueError as e:
            if "does not exist" in str(e):
                raise ValueError(f"Лист '{sheet_name}' не найден в файле '{self.full_path}'. Проверьте название листа.")
            else:
                raise e  # другие ValueError (например, проблемы с парсингом)
        except FileNotFoundError:
            # Это уже обработано выше, но на всякий случай
            raise
        except Exception as e:
            # Все остальные ошибки (включая ошибки pandas) — оборачиваем в RuntimeError
            raise RuntimeError(f"Ошибка при чтении Excel-файла: {e}")

    def read_existing_report(self, file_path: str) -> List[Dict]:
        """
        Читает существующий отчёт и возвращает список данных о НЕ найденных накладных.

        Логика:
        1. Открывает лист "Полный отчёт"
        2. Пропускает строки до заголовков
        3. Фильтрует строки, где Статус == "not_found"
        4. Возвращает список словарей с ключами:
           - number
           - crm_id
           - address
           - isa_amount
           - sfa_amount
           - delivery_city
           - prefix
           - region

        Args:
            file_path (str): Путь к существующему отчёту

        Returns:
            List[Dict]: Только накладные со статусом not_found
        """
        try:
            wb = load_workbook(file_path, read_only=True)
            if "Полный отчёт" not in wb.sheetnames:
                self.logger.error("❌ В отчёте нет листа 'Полный отчёт'")
                return []

            ws = wb["Полный отчёт"]
            records = []

            # === Поиск заголовков ===
            headers = {}
            header_row = None
            for row in ws.iter_rows(min_row=1, max_row=20):
                for cell in row:
                    if cell.value and str(cell.value).strip() == "Номер":
                        header_row = row[0].row
                        break
                if header_row:
                    break

            if not header_row:
                self.logger.error("❌ Не найдены заголовки в отчёте")
                return []

            # === Определяем индексы колонок ===
            for cell in ws[header_row]:
                if cell.value:
                    headers[str(cell.value).strip().lower()] = cell.column - 1

            required = {"номер", "crm id", "адрес", "isa", "sfa", "город", "префикс", "регион", "статус"}
            if not required.issubset(headers.keys()):
                missing = required - set(headers.keys())
                self.logger.error(f"❌ Отсутствуют колонки: {missing}")
                return []

            # === Читаем данные ===
            for row in ws.iter_rows(min_row=header_row + 1):
                if not row[headers["номер"]].value:
                    continue

                status = str(row[headers["статус"]].value).strip().lower()
                if status != "not_found":
                    continue  # пропускаем найденные

                record = {
                    "number": str(row[headers["номер"]].value).strip(),
                    "crm_id": str(row[headers["crm id"]].value).strip() if row[headers["crm id"]].value else None,
                    "address": str(row[headers["адрес"]].value).strip() if row[headers["адрес"]].value else None,
                    "isa_amount": float(row[headers["isa"]].value) if row[headers["isa"]].value else None,
                    "sfa_amount": float(row[headers["sfa"]].value) if row[headers["sfa"]].value else None,
                    "delivery_city": str(row[headers["город"]].value).strip() if row[headers["город"]].value else None,
                    "prefix": str(row[headers["префикс"]].value).strip() if row[headers["префикс"]].value else None,
                    "region": str(row[headers["регион"]].value).strip() if row[headers["регион"]].value else None,
                }
                records.append(record)

            self.logger.info(f"✅ Прочитано {len(records)} накладных из отчёта для обновления")
            return records

        except Exception as e:
            self.logger.error(f"❌ Ошибка при чтении отчёта: {e}")
            return []
        finally:
            if 'wb' in locals():
                wb.close()