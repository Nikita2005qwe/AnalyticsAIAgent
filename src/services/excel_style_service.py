"""
Модуль: excel_style_service.py
Описание: Сервис для работы с цветовым форматированием Excel-файлов.

Назначение:
- Централизованное хранение и применение стилей (заливок) для ячеек.
- Определение цвета заливки на основе статуса накладной и её региона.
- Поддержка единого визуального стиля во всех процессах.

Архитектурная роль:
- Утилита для **форматирования Excel**.
- Не зависит от UI.
- Используется в `InvoiceCheckerProcess._mark_excel()` и других процессах, где нужно подсвечивать результаты.

Зависимости:
- openpyxl.styles.PatternFill — для создания заливок
- models.invoice.Invoice — для анализа статуса и региона
"""

# --- Импорты из стандартной библиотеки ---
from typing import Optional

# --- Импорты из сторонних библиотек ---
from openpyxl.styles import PatternFill

# --- Импорты из моделей ---
from src.models.invoice import Invoice

# Цвета для Excel (могут быть вынесены в настройки в будущем)
FILL_FOUND = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
# 🟩 Зелёный: накладная найдена в DMS

FILL_NOT_FOUND = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
# 🟥 Красный: накладная не найдена

FILL_ERROR = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
# 🟨 Жёлтый: ошибка при проверке (таймаут, исключение)

FILL_UNKNOWN_PREFIX = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
# 🟨 Светло-жёлтый: неизвестный префикс (не удалось определить регион)

FILL_OTHER_REGION = PatternFill(start_color="CCDDFF", end_color="CCDDFF", fill_type="solid")
# 🟦 Светло-синий: накладная из другого региона (например, проверяли Сибирь, а накладная из Урала)

FILL_PENDING = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
# ⬜ Серый: статус "ожидает" (не проверялась)

class ExcelStyleService:
    """
    Сервис для применения цветового форматирования к ячейкам Excel.

    Атрибуты:
    - Нет (сервис без состояния — pure functions)
    """

    def get_fill_for_invoice(self, invoice: Invoice, current_region: Optional[str] = None) -> PatternFill:
        """
        Определяет, какую заливку применить к ячейке с номером накладной.

        Args:
            invoice (Invoice): Объект накладной
            current_region (str, optional): Регион, который сейчас проверяется (например, "siberia")

        Returns:
            PatternFill: Цветовая заливка для ячейки

        Логика выбора:
        1. Если у накладной есть ошибка → FILL_ERROR
        2. Если найдена → FILL_FOUND
        3. Если не найдена → FILL_NOT_FOUND
        4. Если статус "unknown" (неизвестный префикс) → FILL_UNKNOWN_PREFIX
        5. Если найдена, но из другого региона → FILL_OTHER_REGION
        6. Иначе → FILL_PENDING (на всякий случай)

        Примеры:
            - invoice.found_in_dms = True → зелёный
            - invoice.found_in_dms = False → красный
            - invoice.error_message = "Timeout" → жёлтый
            - invoice.region = "ural", current_region = "siberia" → синий

        Используется в:
            InvoiceCheckerProcess._mark_excel()
            Будущих процессах с Excel-отчётами
        """
        pass

    def apply_fill_to_cell(self, cell, invoice: Invoice, current_region: Optional[str] = None):
        """
        Применяет заливку к ячейке на основе накладной.

        Args:
            cell: Ячейка openpyxl (например, ws['A1'])
            invoice (Invoice): Накладная
            current_region (str, optional): Текущий регион проверки

        Шаги:
        1. Получить заливку через self.get_fill_for_invoice()
        2. Установить cell.fill = fill

        Удобно использовать в цикле:
            for cell, invoice in zip(cells, invoices):
                style_service.apply_fill_to_cell(cell, invoice, current_region="siberia")
        """
        pass

    def get_fill_by_status(self, status: str) -> PatternFill:
        """
        Возвращает заливку по строковому статусу.

        Args:
            status (str): "found", "not_found", "error", "unknown", "pending"

        Returns:
            PatternFill: Соответствующая заливка

        Используется, если у тебя нет объекта Invoice, а только статус.
        Например, при восстановлении из лога.
        """
        pass