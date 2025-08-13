# src/buisness_processes/data/settings_for_invoices_check.py

from openpyxl.styles import PatternFill

# Соответствие префикса → город → регион
INVOICE_PREFIX_REGIONS: dict[str, dict] = {
    # Сибирь
    "01/": {"city": "Красноярск", "region": "siberia"},
    "02/": {"city": "Абакан", "region": "siberia"},
    "04/": {"city": "Новокузнецк", "region": "siberia"},
    "05/": {"city": "Новосибирск", "region": "siberia"},
    "Б-": {"city": "Абакан", "region": "siberia"},
    "К-": {"city": "Красноярск", "region": "siberia"},
    "И-": {"city": "Новокузнецк", "region": "siberia"},

    # Урал
    "07/": {"city": "Челябинск", "region": "ural"},
    "Ч-": {"city": "Челябинск", "region": "ural"},
}

# Цвета для Excel
FILL_FOUND = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")      # Зелёный
FILL_NOT_FOUND = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Красный
FILL_NOT_SEARCHED = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # Жёлтый
FILL_OTHER_REGION = PatternFill(start_color="CCDDFF", end_color="CCDDFF", fill_type="solid")  # Светло-синий (не этот регион)