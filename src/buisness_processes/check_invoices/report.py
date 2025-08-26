"""
Модуль: report.py
Описание: Генерация отчёта по результатам проверки накладных.

Назначение:
- Создание Excel-отчёта с двумя листами:
  1. "Не найденные по городам" — только not_found, по столбцам (города)
  2. "Полный отчёт" — все накладные с полной информацией и цветами
- Не зависит от способа получения данных
- Использует openpyxl для форматирования

Зависимости:
- openpyxl
- CheckedInvoice, CheckStatus, Invoice — из модели
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
import os

from src.core.logger import Logger
from src.models.invoice.invoice import CheckedInvoice, CheckStatus


# === Цвета ===
FILL_FOUND = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")      # 🟩
FILL_NOT_FOUND = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # 🟥
FILL_ERROR = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")      # 🟨
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")    # 🔵
FONT_HEADER = Font(color="FFFFFF", bold=True)


class ReporterOfCheckInvoiceProcess:
    """
    Генерирует отчёт по результатам проверки накладных.

    Атрибуты:
    - report_file (str): Путь к файлу отчёта
    - data (List[CheckedInvoice]): Список проверенных накладных
    """

    def __init__(self, file_name: str, data: list[CheckedInvoice], logger: Logger):
        self.report_file = file_name
        self.data = data
        self.logger = logger

    def create_report(self):
        """Создаёт отчёт с двумя листами в нужном формате"""
        wb = Workbook()
        wb.remove(wb.active)  # удаляем пустой лист

        # Лист 1: Не найденные по городам
        self._create_not_found_by_city_sheet(wb)

        # Лист 2: Полный отчёт
        self._create_full_report_sheet(wb)

        # Сохраняем
        try:
            wb.save(self.report_file)
            print(f"✅ Отчёт сохранён: {os.path.abspath(self.report_file)}")
        except Exception as e:
            print(f"❌ Ошибка при сохранении отчёта: {e}")
            raise

    def update_report(self, report_path: str, data: List[CheckedInvoice]):
        """
        Обновляет существующий отчёт:
        - Удаляет старые листы
        - Создаёт новые
        - Сохраняет в тот же файл
        """
        try:
            wb = load_workbook(report_path)

            # Удаляем старые листы
            for sheet_name in ["Не найденные по городам", "Полный отчёт"]:
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]

            # Пересоздаём отчёт
            self.data = data
            self._create_not_found_by_city_sheet(wb)
            self._create_full_report_sheet(wb)

            wb.save(report_path)
            self.logger.info(f"✅ Отчёт обновлён: {report_path}")
            self.open_report()

        except Exception as e:
            self.logger.error(f"❌ Ошибка при обновлении отчёта: {e}")
            raise

    def _create_not_found_by_city_sheet(self, wb):
        """Создаёт лист с накладными, не найденными в DMS, по городам (столбцы)"""
        ws = wb.create_sheet(title="Не найденные по городам")

        # Фильтруем только not_found
        not_found_invoices = [
            ci for ci in self.data
            if ci.status == CheckStatus.NOT_FOUND
        ]

        if not not_found_invoices:
            ws["A1"] = "Нет накладных со статусом 'not_found'"
            return

        # Группируем по городу
        cities = self._group_by_city(not_found_invoices)

        # Заголовки — города
        for col_idx, city in enumerate(sorted(cities.keys()), 1):
            cell = ws.cell(row=1, column=col_idx, value=city)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = Alignment(horizontal="center")

            # Номера накладных
            for row_offset, ci in enumerate(cities[city], 2):
                ws.cell(row=row_offset, column=col_idx, value=ci.invoice.number)

        # Автоподбор ширины
        self._auto_adjust_columns(ws)

    def _create_full_report_sheet(self, wb):
        """Создаёт лист с полным отчётом по всем накладным, сгруппированным по городам"""
        ws = wb.create_sheet(title="Полный отчёт")

        headers = ["Номер", "CRM ID", "Адрес", "ISA", "SFA", "Город", "Префикс", "Регион", "Статус"]
        current_row = 1

        # Группируем все накладные по городу
        all_invoices = [ci for ci in self.data]
        cities = self._group_by_city(all_invoices, key=lambda ci: ci.invoice.delivery_city)

        # Сортируем города
        for city in sorted(cities.keys()):
            # Заголовок города
            cell = ws.cell(row=current_row, column=1, value=city)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
            current_row += 1

            # Заголовки таблицы
            if current_row == 2 or (current_row > 2 and city == sorted(cities.keys())[0]):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                current_row += 1

            # Данные
            for ci in cities[city]:
                row = [
                    ci.invoice.number,
                    ci.invoice.crm_id,
                    ci.invoice.address,
                    ci.invoice.isa_amount,
                    ci.invoice.sfa_amount if ci.invoice.sfa_amount is not None else "",
                    ci.invoice.delivery_city,
                    ci.invoice.prefix,
                    ci.invoice.region,
                    ci.status.value
                ]
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    self._apply_status_fill(cell, ci.status)

                current_row += 1

            # Пустая строка между городами
            current_row += 1

        # Автоподбор ширины
        self._auto_adjust_columns(ws)

        # Перенос текста для адреса
        for cell in ws['C']:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _group_by_city(self, items: List[Any], key=None) -> Dict[str, List]:
        """
        Группирует элементы по городу доставки.

        :param items: Список CheckedInvoice или любых объектов
        :param key: Функция для извлечения города (по умолчанию: ci.invoice.delivery_city)
        :return: Словарь {город: [номера или объекты]}
        """
        if key is None:
            key = lambda ci: ci.invoice.delivery_city

        grouped = {}
        for item in items:
            city = key(item)
            if city not in grouped:
                grouped[city] = []
            if hasattr(item, 'invoice'):
                grouped[city].append(item)
            else:
                grouped[city].append(item)
        return grouped

    def _apply_status_fill(self, cell, status: CheckStatus):
        """Применяет заливку в зависимости от статуса"""
        if status == CheckStatus.FOUND:
            cell.fill = FILL_FOUND
        elif status == CheckStatus.NOT_FOUND:
            cell.fill = FILL_NOT_FOUND
        elif status == CheckStatus.ERROR:
            cell.fill = FILL_ERROR

    def _auto_adjust_columns(self, ws):
        """Автоподбор ширины столбцов"""
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    def open_report(self):
        """Открывает отчётный файл системной командой."""
        try:
            abs_path = os.path.abspath(self.report_file)
            if not os.path.exists(abs_path):
                self.logger.error(f"❌ Файл отчёта не найден: {abs_path}")
                return

            if os.name == "nt":  # Windows
                os.startfile(abs_path)
            elif os.name == "posix":  # macOS, Linux
                import subprocess
                if "darwin" in os.uname().sysname:  # macOS
                    subprocess.run(["open", abs_path])
                else:  # Linux
                    subprocess.run(["xdg-open", abs_path])
            print(f"📂 Отчёт открыт: {abs_path}")
        except Exception as e:
            print(f"⚠️ Не удалось открыть отчёт: {e}")