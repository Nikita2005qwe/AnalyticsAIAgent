# src/buisness_processes/check_invoices/process.py

from typing import List, Tuple, Dict
import pandas as pd
from openpyxl import load_workbook

from src.buisness_processes.data.settings_for_invoices_check import (
    INVOICE_PREFIX_REGIONS,
    FILL_FOUND,
    FILL_NOT_FOUND,
    FILL_NOT_SEARCHED,
    FILL_OTHER_REGION,
)
from src.buisness_processes.check_invoices.operation import DMSOperation


class InvoiceCheckerProcess:
    def __init__(self, log_callback):
        self.log = log_callback

    def run(self, file_path: str, sheet_name: str) -> List[Tuple[str, bool, str]]:
        """
        Проверяет накладные, разделяя по регионам.
        Сначала — Сибирь, потом — Урал.
        """
        # Загружаем номера
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, usecols="A", header=0, names=["invoice"])
        except Exception as e:
            self.log(f"❌ Ошибка чтения Excel: {e}")
            return []

        numbers = df["invoice"].dropna().astype(str).str.strip().tolist()
        if not numbers:
            self.log("🟡 Файл пуст.")
            return []

        # Разделяем по регионам
        grouped: Dict[str, List[str]] = {"siberia": [], "ural": []}
        unknown = []

        for num in numbers:
            if not num or num.lower() == "nan":
                continue
            region = self._get_region(num)
            if region in grouped:
                grouped[region].append(num)
            else:
                unknown.append(num)

        # Результаты
        all_results = []

        # --- Обработка: Сибирь ---
        if grouped["siberia"]:
            self.log("🌍 Проверка накладных: Сибирь")
            op = DMSOperation("siberia")
            for num in grouped["siberia"]:
                try:
                    found = op.check_invoice(num)
                    all_results.append((num, found, "found" if found else "not_found"))
                    self.log(f"   → {num}: {'✅' if found else '❌'}")
                except Exception as e:
                    self.log(f"   ⚠️ {num}: {e}")
                    all_results.append((num, False, "error"))
            op.close()

        # --- Обработка: Урал ---
        if grouped["ural"]:
            self.log("🌍 Проверка накладных: Урал")
            op = DMSOperation("ural")
            for num in grouped["ural"]:
                try:
                    found = op.check_invoice(num)
                    all_results.append((num, found, "found" if found else "not_found"))
                    self.log(f"   → {num}: {'✅' if found else '❌'}")
                except Exception as e:
                    self.log(f"   ⚠️ {num}: {e}")
                    all_results.append((num, False, "error"))
            op.close()

        # --- Неизвестные ---
        for num in unknown:
            all_results.append((num, False, "unknown"))
            self.log(f"⚠️ Неизвестный префикс: {num}")

        # --- Обновляем Excel ---
        self._mark_excel(file_path, sheet_name, all_results)
        return all_results

    def _get_region(self, invoice_number: str) -> str:
        """Возвращает регион по префиксу."""
        for prefix, data in INVOICE_PREFIX_REGIONS.items():
            if invoice_number.startswith(prefix):
                return data["region"]
        return "unknown"

    def _mark_excel(self, file_path: str, sheet_name: str, results: List[Tuple[str, bool, str]]):
        """Подсвечивает ячейки в Excel по статусу."""
        try:
            wb = load_workbook(file_path)
            ws = wb[sheet_name]
            status_map = {num: status for num, _, status in results}

            row = 2
            while True:
                cell = ws[f"A{row}"]
                if cell.value is None:
                    break
                num = str(cell.value).strip()
                status = status_map.get(num, "unknown")

                if status == "found":
                    cell.fill = FILL_FOUND
                elif status == "not_found":
                    cell.fill = FILL_NOT_FOUND
                elif status == "unknown":
                    cell.fill = FILL_NOT_SEARCHED
                else:  # error / other
                    cell.fill = FILL_OTHER_REGION

                row += 1

            wb.save(file_path)
            self.log(f"🎨 Результаты сохранены: {file_path}")
        except Exception as e:
            self.log(f"⚠️ Ошибка при обновлении Excel: {e}")